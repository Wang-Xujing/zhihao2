from django.contrib import admin
from apps import *
from functools import partial, reduce, update_wrapper
from django.views.generic import RedirectView
from apps.e_breed.models import *
from apps.e_breed.models import LambInfo

IS_POPUP_VAR = '_popup'
from django.http import HttpResponse
from django.utils.translation import gettext as _, gettext_lazy

from openpyxl import Workbook
from django.utils.datastructures import MultiValueDictKeyError
import copy
import json
import operator
import re
from functools import partial, reduce, update_wrapper
from urllib.parse import quote as urlquote
from django.contrib.admin import actions as ac
from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.admin import helpers, widgets
from django.contrib.admin.checks import (
    BaseModelAdminChecks, InlineModelAdminChecks, ModelAdminChecks,
)
from django.contrib.admin.exceptions import DisallowedModelAdminToField
from django.contrib.admin.templatetags.admin_urls import add_preserved_filters
from django.contrib.admin.utils import (
    NestedObjects, construct_change_message, flatten_fieldsets,
    get_deleted_objects, lookup_needs_distinct, model_format_dict,
    model_ngettext, quote, unquote,
)
from django.contrib.admin.views.autocomplete import AutocompleteJsonView
from django.contrib.admin.widgets import (
    AutocompleteSelect, AutocompleteSelectMultiple,
)
from django.contrib.auth import get_permission_codename
from django.core.exceptions import (
    FieldDoesNotExist, FieldError, PermissionDenied, ValidationError,
)
from django.core.paginator import Paginator
from django.db import models, router, transaction
from django.db.models.constants import LOOKUP_SEP
from django.db.models.fields import BLANK_CHOICE_DASH
from django.forms.formsets import DELETION_FIELD_NAME, all_valid
from django.forms.models import (
    BaseInlineFormSet, inlineformset_factory, modelform_defines_fields,
    modelform_factory, modelformset_factory,
)
from django.forms.widgets import CheckboxSelectMultiple, SelectMultiple
from django.http import HttpResponseRedirect
from django.http.response import HttpResponseBase
from django.template.response import SimpleTemplateResponse, TemplateResponse
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.utils.html import format_html
from django.utils.http import urlencode
from django.utils.safestring import mark_safe
from django.utils.text import capfirst, format_lazy, get_text_list
from django.utils.translation import gettext as _, ngettext
from django.views.decorators.csrf import csrf_protect
from django.views.generic import RedirectView


class IncorrectLookupParameters(Exception):
    pass


@admin.register(rutInfo)
class rutInfoAdmin(admin.ModelAdmin):
    field_name = rutInfo()._meta.fields
    list_display = ['id', 'ele_id', 'pre_num']

    class eleFilter(admin.SimpleListFilter):
        title = "电子耳号"
        parameter_name = 'basic_id'

        def lookups(self, request, model_admin):
            params = rutInfo.objects.filter(belong=request.user.belong).values('basic_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['basic_id']).values("ele_num").first()
                temp = ('{}'.format(basic_ele['ele_num']), ('{}'.format(basic_ele['ele_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(ele_num=value).values("id").first()
            return rutInfo.objects.filter(basic_id=basic['id']).all()

    class preFilter(admin.SimpleListFilter):
        title = "防疫耳号"
        parameter_name = 'pre_num'

        def lookups(self, request, model_admin):
            params = rutInfo.objects.filter(belong=request.user.belong).values('basic_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['basic_id']).values("pre_num").first()
                temp = ('{}'.format(basic_ele['pre_num']), ('{}'.format(basic_ele['pre_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(pre_num=value).values("id").first()
            return rutInfo.objects.filter(basic_id=basic['id']).all()

    list_filter =[]
    exclude_display = ['belong', 'basic_id']
    exclude_filter = []
    search_fields = []

    doit(field_name, list_display, exclude_display, list_filter, exclude_filter, search_fields)
    ordering = ('-f_date',)
    change_list_template = 'change_list_zhu.html'
    add_form_template = 'test-rutinfo.html'
    change_list_template_page = 'test_change_list.html'


@admin.register(SemencollectInfo)
class SemencollectInfoAdmin(admin.ModelAdmin):
    field_name = SemencollectInfo()._meta.fields
    list_display = ['id', 'ele_id', 'pre_num']

    class eleFilter(admin.SimpleListFilter):
        title = "电子耳号"
        parameter_name = 'basic_id'

        def lookups(self, request, model_admin):
            params = SemencollectInfo.objects.filter(belong=request.user.belong).values('basic_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['basic_id']).values("ele_num").first()
                temp = ('{}'.format(basic_ele['ele_num']), ('{}'.format(basic_ele['ele_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(ele_num=value).values("id").first()
            return SemencollectInfo.objects.filter(basic_id=basic['id']).all()

    class preFilter(admin.SimpleListFilter):
        title = "防疫耳号"
        parameter_name = 'pre_num'

        def lookups(self, request, model_admin):
            params = SemencollectInfo.objects.filter(belong=request.user.belong).values('basic_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['basic_id']).values("pre_num").first()
                temp = ('{}'.format(basic_ele['pre_num']), ('{}'.format(basic_ele['pre_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(pre_num=value).values("id").first()
            return SemencollectInfo.objects.filter(basic_id=basic['id']).all()

    list_filter =[]
    exclude_display = ['belong']
    exclude_filter = []
    search_fields = []

    doit(field_name, list_display, exclude_display, list_filter, exclude_filter, search_fields)
    ordering = ('-f_date',)
    change_list_template = 'change_list_zhu.html'
    add_form_template = 'test-semencollectinfo.html'
    change_list_template_page = 'test_change_list.html'


@admin.register(breedingInfo)
class breedingInfoAdmin(admin.ModelAdmin):
    def __init__(self, model, admin_site):
        super().__init__(breedingInfo, admin_site)
        self.model = breedingInfo
        self.to_zhong = {
        }
        self.basic_fields = ['pre_num', 'ele_num', 'variety', 'sex', 'purpose', 'state', 'house_name', 'hurdle_name',
                             'birth', 'bir_weight', 'mon_age']
        self.pregnant_fields = ['check_type', 'In_pregnancy']
        self.postnatal_fields = ['delivery_date', 'ewe_health', 'live_num', 'Booroola']
        self.lamb_fields = ['ele_num', 'variety', 'sex', 'bir_weight']

    field_name = breedingInfo()._meta.fields
    actions_selection_counter = False
    list_display = ['ele_id_ewe', 'pre_num_ewe', 'ele_id_ram', 'pre_num_ram']

    class eweFilter(admin.SimpleListFilter):

        title = "母羊耳号"
        parameter_name = 'ewe_id'

        def lookups(self, request, model_admin):
            params = breedingInfo.objects.filter(belong=request.user.belong).values('ewe_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['ewe_id']).values("ele_num").first()
                temp = ('{}'.format(basic_ele['ele_num']), ('{}'.format(basic_ele['ele_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(ele_num=value).values("id").first()
            return breedingInfo.objects.filter(ewe_id=basic['id']).all()

    class ramFilter(admin.SimpleListFilter):

        title = "公羊耳号"
        parameter_name = 'ram_id'

        def lookups(self, request, model_admin):
            params = breedingInfo.objects.filter(belong=request.user.belong).values('ram_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['ram_id']).values("ele_num").first()
                temp = ('{}'.format(basic_ele['ele_num']), ('{}'.format(basic_ele['ele_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(ele_num=value).values("id").first()
            return breedingInfo.objects.filter(ram_id=basic['id']).all()

    class ewepreFilter(admin.SimpleListFilter):
        title = "防疫耳号"
        parameter_name = 'ewepre_num'

        def lookups(self, request, model_admin):
            params = breedingInfo.objects.filter(belong=request.user.belong).values('ewe_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['ewe_id']).values("pre_num").first()
                temp = ('{}'.format(basic_ele['pre_num']), ('{}'.format(basic_ele['pre_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(pre_num=value).values("id").first()
            return breedingInfo.objects.filter(ewe_id=basic['id']).all()

    class rampreFilter(admin.SimpleListFilter):
        title = "防疫耳号"
        parameter_name = 'rampre_num'

        def lookups(self, request, model_admin):
            params = breedingInfo.objects.filter(belong=request.user.belong).values('ram_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['ram_id']).values("pre_num").first()
                temp = ('{}'.format(basic_ele['pre_num']), ('{}'.format(basic_ele['pre_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(pre_num=value).values("id").first()
            return breedingInfo.objects.filter(ram_id=basic['id']).all()

    # list_filter = [eweFilter, ewepreFilter, ramFilter, rampreFilter]
    list_filter = []
    exclude_display = ['belong', 'ewe_id', 'ram_id']
    exclude_filter = []
    search_fields = []

    doit(field_name, list_display, exclude_display, list_filter, exclude_filter, search_fields)
    ordering = ('-f_date',)
    change_list_template = 'change_list_zhu.html'
    add_form_template = 'e_breed/test_breedinginfo.html'
    change_form_template = 'e_breed/test_breedinginfo.html'
    change_list_template_page = 'test_change_list.html'

    def changelist_view(self, request, extra_context=None):
        """
        The 'change list' admin view for this model.
        """

        ##解除只能选中执行的限制，增加是否执行act的bool
        act_do = request.POST['action'] == 'exportAll_to_excel' if 'action' in request.POST else False

        from django.contrib.admin.views.main import ERROR_FLAG
        opts = self.model._meta
        app_label = opts.app_label
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        try:
            cl = self.get_changelist_instance(request)
        except IncorrectLookupParameters:
            # Wacky lookup parameters were given, so redirect to the main
            # changelist page, without parameters, and pass an 'invalid=1'
            # parameter via the query string. If wacky parameters were given
            # and the 'invalid=1' parameter was already in the query string,
            # something is screwed up with the database, so display an error
            # page.
            if ERROR_FLAG in request.GET:
                return SimpleTemplateResponse('admin/invalid_setup.html', {
                    'title': _('Database error'),
                })
            return HttpResponseRedirect(request.path + '?' + ERROR_FLAG + '=1')

        # If the request was POSTed, this might be a bulk action or a bulk
        # edit. Try to look up an action or confirmation first, but if this
        # isn't an action the POST will fall through to the bulk edit check,
        # below.
        action_failed = False
        selected = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)

        actions = self.get_actions(request)
        # Actions with no confirmation
        if (actions and request.method == 'POST' and
                'index' in request.POST and '_save' not in request.POST):
            if selected or act_do:
                response = self.response_action(request, queryset=cl.get_queryset(request))
                if response:
                    return response
                else:
                    action_failed = True
            else:
                msg = _("Items must be selected in order to perform "
                        "actions on them. No items have been changed.")
                self.message_user(request, msg, messages.WARNING)
                action_failed = True

        if (actions and request.method == 'POST' and
            helpers.ACTION_CHECKBOX_NAME in request.POST and
            'index' not in request.POST and '_save' not in request.POST) or act_do:
            if selected or act_do:
                response = self.response_action(request, queryset=cl.get_queryset(request))
                if response:
                    return response
                else:
                    action_failed = True

        if action_failed:
            # Redirect back to the changelist page to avoid resubmitting the
            # form if the user refreshes the browser or uses the "No, take
            # me back" button on the action confirmation page.
            return HttpResponseRedirect(request.get_full_path())

        # If we're allowing changelist editing, we need to construct a formset
        # for the changelist given all the fields to be edited. Then we'll
        # use the formset to validate/process POSTed data.
        formset = cl.formset = None

        # Handle POSTed bulk-edit data.
        if request.method == 'POST' and cl.list_editable and '_save' in request.POST:
            if not self.has_change_permission(request):
                raise PermissionDenied
            FormSet = self.get_changelist_formset(request)
            modified_objects = self._get_list_editable_queryset(request, FormSet.get_default_prefix())
            formset = cl.formset = FormSet(request.POST, request.FILES, queryset=modified_objects)
            if formset.is_valid():
                changecount = 0
                for form in formset.forms:
                    if form.has_changed():
                        obj = self.save_form(request, form, change=True)
                        self.save_model(request, obj, form, change=True)
                        self.save_related(request, form, formsets=[], change=True)
                        change_msg = self.construct_change_message(request, form, None)
                        self.log_change(request, obj, change_msg)
                        changecount += 1

                if changecount:
                    msg = ngettext(
                        "%(count)s %(name)s was changed successfully.",
                        "%(count)s %(name)s were changed successfully.",
                        changecount
                    ) % {
                              'count': changecount,
                              'name': model_ngettext(opts, changecount),
                          }
                    self.message_user(request, msg, messages.SUCCESS)

                return HttpResponseRedirect(request.get_full_path())

        # Handle GET -- construct a formset for display.
        elif cl.list_editable and self.has_change_permission(request):
            FormSet = self.get_changelist_formset(request)
            formset = cl.formset = FormSet(queryset=cl.result_list)

        # Build the list of media to be used by the formset.
        if formset:
            media = self.media + formset.media
        else:
            media = self.media

        # Build the action form and populate it with available actions.
        model_name = '_'.join(str(cl.opts).split('.'))
        if actions:
            action_form = self.action_form(auto_id=None)
            action_form.fields['action'].choices = self.get_action_choices_no_delete(request, model_name)
            media += action_form.media
        else:
            action_form = None

        selection_note_all = ngettext(
            '%(total_count)s selected',
            'All %(total_count)s selected',
            cl.result_count
        )
        context = {
            **self.admin_site.each_context(request),
            'module_name': str(opts.verbose_name_plural),
            'selection_note': _('0 of %(cnt)s selected') % {'cnt': len(cl.result_list)},
            'selection_note_all': selection_note_all % {'total_count': cl.result_count},
            'title': cl.title,
            'is_popup': cl.is_popup,
            'to_field': cl.to_field,
            'model_name': model_name,
            'cl': cl,
            'media': media,
            'has_add_permission': self.has_add_permission(request),
            'opts': cl.opts,
            'action_form': action_form,
            'actions_on_top': self.actions_on_top,
            'actions_on_bottom': self.actions_on_bottom,
            'actions_selection_counter': self.actions_selection_counter,
            'preserved_filters': self.get_preserved_filters(request),
            **(extra_context or {}),
        }

        request.current_app = self.admin_site.name

        return TemplateResponse(request, self.change_list_template or [
            'admin/%s/%s/change_list.html' % (app_label, opts.model_name),
            'admin/%s/change_list.html' % app_label,
            'admin/change_list.html'
        ], context)

    def response_action(self, request, queryset):
        """
        Handle an admin action. This is called if a request is POSTed to the
        changelist; it returns an HttpResponse if the action was handled, and
        None otherwise.
        """
        act_do = request.POST['action'] == 'exportAll_to_excel' if 'action' in request.POST else False
        # There can be multiple action forms on the page (at the top
        # and bottom of the change list, for example). Get the action
        # whose button was pushed.

        try:
            action_index = int(request.POST.get('index', 0))
        except ValueError:
            action_index = 0

        # Construct the action form.
        data = request.POST.copy()
        data.pop(helpers.ACTION_CHECKBOX_NAME, None)
        data.pop("index", None)

        # Use the action whose button was pushed
        try:
            data.update({'action': data.getlist('action')[action_index]})
        except IndexError:
            # If we didn't get an action from the chosen form that's invalid
            # POST data, so by deleting action it'll fail the validation check
            # below. So no need to do anything here
            pass

        action_form = self.action_form(data, auto_id=None)
        action_form.fields['action'].choices = self.get_action_choices(request)

        # If the form's valid we can handle the action.
        if action_form.is_valid():
            action = action_form.cleaned_data['action']
            select_across = action_form.cleaned_data['select_across']
            func = self.get_actions(request)[action][0]

            # Get the list of selected PKs. If nothing's selected, we can't
            # perform an action on it, so bail. Except we want to perform
            # the action explicitly on all objects.
            selected = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)
            if not selected and not select_across and not act_do:
                # Reminder that something needs to be selected or nothing will happen
                msg = _("Items must be selected in order to perform "
                        "actions on them. No items have been changed.")
                self.message_user(request, msg, messages.WARNING)
                return None

            if not select_across:
                # Perform the action only on the selected objects
                queryset = queryset.filter(pk__in=selected)

            response = func(self, request, queryset)

            # Actions may return an HttpResponse-like object, which will be
            # used as the response from the POST. If not, we'll be a good
            # little HTTP citizen and redirect back to the changelist page.
            if isinstance(response, HttpResponseBase):
                return response
            else:
                return HttpResponseRedirect(request.get_full_path())
        else:
            msg = _("No action selected.")
            self.message_user(request, msg, messages.WARNING)
            return None

    # 检验这个字段是否是choice字段，如果是就需要将其对应起来
    def check_choices(self, models):
        data = {}
        for i in models:
            fields = i._meta.fields
            for j in fields:
                if j.choices:
                    data[j.name] = dict(j.choices)
        return data

    def excel_header(self, fields_name):
        dict = self.to_zhong
        data = []
        # basic=[i.verbose_name for i in BasicInfo._meta.fields if i.name in self.basic_fields]# 导入基本信息字段名

        breed = [dict[i.name] if dict.get(i.name) else i.verbose_name for i in self.model._meta.fields if
                 i.name in fields_name]

        basic_all = {}
        basic = []
        pregnant_all = {}
        pregnant = []
        postnata_all = {}
        postnata = []
        lamb_all = {}
        lamb = []
        for i in BasicInfo._meta.fields:
            basic_all[i.name] = i.verbose_name
        for i in self.basic_fields:
            basic.append(basic_all[i])

        for i in pregnantInfo._meta.fields:
            pregnant_all[i.name] = i.verbose_name
        for i in self.pregnant_fields:
            pregnant.append(pregnant_all[i])

        for i in postnatalInfo._meta.fields:
            postnata_all[i.name] = i.verbose_name
        for i in self.postnatal_fields:
            postnata.append(postnata_all[i])

        for i in LambInfo._meta.fields:
            lamb_all[i.name] = i.verbose_name
        for j in range(1, 6):
            for i in self.lamb_fields:
                lamb.append('后裔' + str(j) + lamb_all[i])
        basic.extend(['父电子耳号', '母电子耳号', '祖父电子耳号', '祖母电子耳号', '外祖父电子耳号', '外祖母电子耳号'])
        data.extend(basic)
        data.extend(['母防疫耳号', '父防疫耳号'])
        data.extend(breed)
        data.extend(pregnant)
        data.extend(postnata)
        data.extend(lamb)
        return data

    def add_basic(self, data, id):
        basic = []
        for i in self.basic_fields:
            file = BasicInfo.objects.filter(id=id).values(i).first()
            if i in self.choice_field.keys() and file:
                basic.append(self.choice_field[i][file[i]] if file[i] in self.choice_field[i].keys() else file[i])
            else:
                basic.append(file[i] if file else '')
        f_id = BasicInfo.objects.filter(id=id).values('father_id').first()['father_id'] if BasicInfo.objects.filter(
            id=id).values('father_id').first() else 0
        m_id = BasicInfo.objects.filter(id=id).values('mother_id').first()['mother_id'] if BasicInfo.objects.filter(
            id=id).values('mother_id') else 0
        ff_id = BasicInfo.objects.filter(id=f_id).values('father_id').first()['father_id'] if BasicInfo.objects.filter(
            id=f_id).values('father_id').first() else 0
        fm_id = BasicInfo.objects.filter(id=f_id).values('mother_id').first()['mother_id'] if BasicInfo.objects.filter(
            id=f_id).values('mother_id').first() else 0
        print(BasicInfo.objects.filter(id=f_id).values('father_id').first())
        mf_id = BasicInfo.objects.filter(id=m_id).values('father_id').first()['father_id'] if BasicInfo.objects.filter(
            id=m_id).values('father_id').first() else 0
        mm_id = BasicInfo.objects.filter(id=m_id).values('mother_id').first()['mother_id'] if BasicInfo.objects.filter(
            id=m_id).values('mother_id').first() else 0
        basic.extend([
            BasicInfo.objects.filter(id=f_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=f_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=m_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=m_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=ff_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=ff_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=fm_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=fm_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=mf_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=mf_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=mm_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=mm_id).values('ele_num').first() else '无',
        ])  ##添加爸爸妈妈爷爷奶奶姥姥姥爷

        return basic

    def add_breed(self, obj, fields_name):
        add = []
        # 寻找配种公母羊
        ewe_id = obj.ewe_id
        ram_id = obj.ram_id
        ewe = BasicInfo.objects.filter(id=ewe_id).first()
        ram = BasicInfo.objects.filter(id=ram_id).first()
        add.append(ewe.pre_num if ewe else '')
        add.append(ram.pre_num if ram else '')
        for i in fields_name:
            file = eval('obj.' + i)
            if i in self.choice_field.keys():
                file = self.choice_field[i][file]
            add.append(file)

        return add

    def add_pregnant(self, data, id):
        basic = []
        for i in self.pregnant_fields:
            file = pregnantInfo.objects.filter(breeding_id=id).values(i).first()
            if i in self.choice_field.keys() and file:
                basic.append(self.choice_field[i][file[i]] if file[i] in self.choice_field[i].keys() else file[i])
            else:
                basic.append(file[i] if file else '')

        return basic

    def add_postnatal(self, data, id):
        basic = []
        for i in self.postnatal_fields:
            file = postnatalInfo.objects.filter(breeding_id=id).values(i).first()
            if i in self.choice_field.keys() and file:
                print(file[i])
                if file[i] == 'TRUE':
                    basic.append('正常')
                elif file[i] == 'FALSE':
                    basic.append('不正常')
                else:
                    basic.append(self.choice_field[i][file[i]] if file[i] in self.choice_field[i].keys() else file[i])
            else:
                basic.append(file[i] if file else '')

        return basic

    def add_lamb(self, data, id):
        basic = []
        ids = LambInfo.objects.filter(breeding_id=id).values('id')
        for l_id in ids:
            for i in self.lamb_fields:
                file = LambInfo.objects.filter(id=l_id['id']).values(i).first()
                if i in self.choice_field.keys() and file:
                    print('yes')
                    print(file)
                    basic.append(self.choice_field[i][file[i]] if file[i] in self.choice_field[i].keys() else file[i])
                else:
                    basic.append(file[i] if file else '')
        print(basic)
        return basic

    def data(self, obj, fields_name):
        self.choice_field = self.check_choices([breedingInfo(), BasicInfo(), postnatalInfo])
        data = []
        id = getattr(obj, 'id')
        ewe_id = getattr(obj, 'ewe_id')
        ram_id = getattr(obj, 'ram_id')
        data.extend(self.add_basic(data, ewe_id))  ##添加基本信息
        data.extend(self.add_breed(obj, fields_name))
        data.extend(self.add_pregnant(data, id))
        data.extend(self.add_postnatal(data, id))
        lamb = self.add_lamb(data, id)
        data.extend(lamb if lamb else ['', ])

        # for field in fields_name:
        #     attr=getattr(obj, field)
        #     data.append(attr)
        return data

    def export_to_excel(self, request, queryset):
        fields_name = [i.name for i in self.model._meta.fields]
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response[
            'Content-Disposition'] = f'attachment; filename={self.model._meta.verbose_name.title()}.xlsx'.encode(
            'utf-8')  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        dict = self.to_zhong

        fields_name.remove('id')  ##删除id字段
        fields_name.remove('ewe_id')
        fields_name.remove('ram_id')
        fields_name.remove('belong')
        new_fields_name = self.excel_header(fields_name)

        ws.append(new_fields_name)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            data = self.data(obj, fields_name)
            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response

    def exportAll_to_excel(self, request, queryset):
        queryset = self.model.objects.all()

        fields_name = [i.name for i in self.model._meta.fields]

        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response[
            'Content-Disposition'] = f'attachment; filename={self.model._meta.verbose_name.title()}.xlsx'.encode(
            'utf-8')  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        dict = self.to_zhong

        fields_name.remove('id')  ##删除id字段
        fields_name.remove('ewe_id')
        fields_name.remove('ram_id')
        fields_name.remove('belong')
        new_fields_name = self.excel_header(fields_name)

        ws.append(new_fields_name)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            data = self.data(obj, fields_name)

            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response

    export_to_excel.allowed_permissions = ('export',)
    export_to_excel.short_description = gettext_lazy("导出所选数据")
    exportAll_to_excel.allowed_permissions = ('exportAll',)
    exportAll_to_excel.short_description = gettext_lazy("导出所有数据")

    actions = [export_to_excel, exportAll_to_excel]


@admin.register(pregnantInfo)
class pregnantInfoAdmin(admin.ModelAdmin):
    field_name = pregnantInfo()._meta.fields
    list_display = ['id', 'ele_num', 'pre_num']

    list_filter = []
    # list_filter = [eleFilter,preFilter]
    exclude_display = ['belong']
    exclude_filter = []
    search_fields = []

    doit(field_name, list_display, exclude_display, list_filter, exclude_filter, search_fields)
    list_display.remove('breeding_id')
    ordering = ('-f_date',)
    change_list_template = 'change_list_zhu.html'
    add_form_template = 'test.html'
    change_list_template_page = 'test_change_list.html'


@admin.register(postnatalInfo)
class postnatalInfoAdmin(admin.ModelAdmin):
    field_name = postnatalInfo()._meta.fields
    list_display = ['id', 'ele_id_ewe', 'pre_num_ewe', 'ele_id_ram', 'pre_num_ram']

    class eweFilter(admin.SimpleListFilter):

        title = "母羊耳号"
        parameter_name = 'ewe_id'

        def lookups(self, request, model_admin):
            params = postnatalInfo.objects.filter(belong=request.user.belong).values('ewe_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['ewe_id']).values("ele_num").first()
                temp = ('{}'.format(basic_ele['ele_num']), ('{}'.format(basic_ele['ele_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(ele_num=value).values("id").first()
            return postnatalInfo.objects.filter(ewe_id=basic['id']).all()

    class ramFilter(admin.SimpleListFilter):

        title = "公羊耳号"
        parameter_name = 'ram_id'

        def lookups(self, request, model_admin):
            params = postnatalInfo.objects.filter(belong=request.user.belong).values('ram_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['ram_id']).values("ele_num").first()
                temp = ('{}'.format(basic_ele['ele_num']), ('{}'.format(basic_ele['ele_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(ele_num=value).values("id").first()
            return postnatalInfo.objects.filter(ram_id=basic['id']).all()

    class ewepreFilter(admin.SimpleListFilter):
        title = "防疫耳号"
        parameter_name = 'ewepre_num'

        def lookups(self, request, model_admin):
            params = postnatalInfo.objects.filter(belong=request.user.belong).values('ewe_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['ewe_id']).values("pre_num").first()
                temp = ('{}'.format(basic_ele['pre_num']), ('{}'.format(basic_ele['pre_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(pre_num=value).values("id").first()
            return postnatalInfo.objects.filter(ewe_id=basic['id']).all()

    class rampreFilter(admin.SimpleListFilter):
        title = "防疫耳号"
        parameter_name = 'rampre_num'

        def lookups(self, request, model_admin):
            params = postnatalInfo.objects.filter(belong=request.user.belong).values('ram_id')
            look_choice = []
            for i in list(params):
                basic_ele = BasicInfo.objects.filter(id=i['ram_id']).values("pre_num").first()
                temp = ('{}'.format(basic_ele['pre_num']), ('{}'.format(basic_ele['pre_num'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            basic = BasicInfo.objects.filter(pre_num=value).values("id").first()
            return postnatalInfo.objects.filter(ram_id=basic['id']).all()

    # list_filter = [eweFilter, ewepreFilter, ramFilter, rampreFilter]
    list_filter = []
    exclude_display = ['belong', 'ewe_id', 'ram_id', 'breeding_id']
    exclude_filter = []
    search_fields = []

    doit(field_name, list_display, exclude_display, list_filter, exclude_filter, search_fields)
    ordering = ('-f_date',)
    change_list_template = 'change_list_zhu.html'
    add_form_template = 'test.html'
    change_list_template_page = 'test_change_list.html'


@admin.register(ArtificialfeedingInfo)
class ArtificialfeedingInfoAdmin(admin.ModelAdmin):
    field_name = ArtificialfeedingInfo()._meta.fields
    list_display = ['lamb', 'ele_num', 'pre_num']

    class logoFilter(admin.SimpleListFilter):
        title = "羔羊标识"
        parameter_name = 'lamb_id'

        def lookups(self, request, model_admin):
            params = ArtificialfeedingInfo.objects.filter(belong=request.user.belong).values('lamb_id')
            look_choice = []
            for i in list(params):
                lamb_logo = LambInfo.objects.filter(id=i['lamb_id']).values("logo").first()
                temp = ('{}'.format(lamb_logo['logo']), ('{}'.format(lamb_logo['logo'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            lamb = LambInfo.objects.filter(logo=value).values("id").first()
            return ArtificialfeedingInfo.objects.filter(lamb_id=lamb['id']).all()

    # list_filter = [logoFilter]
    list_filter = []
    exclude_display = ['belong', 'lamb_id']
    exclude_filter = []
    search_fields = []

    doit(field_name, list_display, exclude_display, list_filter, exclude_filter, search_fields)
    ordering = ('-id',)
    change_list_template = 'change_list_zhu.html'
    add_form_template = 'test.html'
    change_list_template_page = 'test_change_list.html'


@admin.register(weaningInfo)
class weaningInfoAdmin(admin.ModelAdmin):
    field_name = weaningInfo()._meta.fields
    list_display = ['lamb', 'ele_num', 'pre_num']

    class logoFilter(admin.SimpleListFilter):
        title = "羔羊标识"
        parameter_name = 'lamb_id'

        def lookups(self, request, model_admin):
            params = weaningInfo.objects.filter(belong=request.user.belong).values('lamb_id')
            look_choice = []
            for i in list(params):
                lamb_logo = LambInfo.objects.filter(id=i['lamb_id']).values("logo").first()
                temp = ('{}'.format(lamb_logo['logo']), ('{}'.format(lamb_logo['logo'])))
                if temp not in look_choice:
                    look_choice.append(temp)
            return look_choice

        def queryset(self, request, queryset):
            # basic_ele = BasicInfo.objects.filter(id=i.basic_id).values("ele_num")
            value = self.value()
            if not value:
                return queryset
            lamb = LambInfo.objects.filter(logo=value).values("id").first()
            return weaningInfo.objects.filter(lamb_id=lamb['id']).all()

    # list_filter = [logoFilter]
    list_filter = []
    exclude_display = ['belong', 'lamb_id']
    exclude_filter = []
    search_fields = []

    doit(field_name, list_display, exclude_display, list_filter, exclude_filter, search_fields)
    ordering = ('-id',)
    change_list_template = 'change_list_zhu.html'
    add_form_template = 'test.html'
    change_list_template_page = 'test_change_list.html'


@admin.register(LambInfo)
class LambInfoAdmin(admin.ModelAdmin):
    field_name = LambInfo()._meta.fields
    list_display = ['id', 'logo']

    list_filter = ['tobasic']
    exclude_display = ['belong', 'id', 'logo', 'father_id', 'mother_id', 'house_id', 'hurdle_id', 'basic_id',
                       'manu_info_id']
    exclude_filter = []
    search_fields = []

    doit(field_name, list_display, exclude_display, list_filter, exclude_filter, search_fields)
    del list_display[0]
    list_display.remove('breeding_id')
    ordering = ('-f_date',)
    change_list_template = 'e_breed/change_list_lamb.html'
    add_form_template = 'test.html'
    change_form_template = 'e_breed/test_change_lamb.html'
    change_list_template_page = 'test_change_list.html'

    def get_urls(self):
        from django.urls import path

        def wrap(view):
            def wrapper(*args, **kwargs):
                return self.admin_site.admin_view(view)(*args, **kwargs)

            wrapper.model_admin = self
            return update_wrapper(wrapper, view)

        info = self.model._meta.app_label, self.model._meta.model_name

        return [
            path('', wrap(self.changelist_view), name='%s_%s_changelist' % info),
            path('add/', wrap(self.add_view), name='%s_%s_add' % info),
            path('page/', wrap(self.page_view), name='%s_%s_page' % info),
            path('lambTobasic/', self.lambTobasic, name='%s_%s_page' % info),
            path('autocomplete/', wrap(self.autocomplete_view), name='%s_%s_autocomplete' % info),
            path('<path:object_id>/history/', wrap(self.history_view), name='%s_%s_history' % info),
            path('<path:object_id>/delete/', wrap(self.delete_view), name='%s_%s_delete' % info),
            path('<path:object_id>/change/', wrap(self.change_view), name='%s_%s_change' % info),
            # For backwards compatibility (was the change url before 1.9)
            path('<path:object_id>/', wrap(RedirectView.as_view(
                pattern_name='%s:%s_%s_change' % ((self.admin_site.name,) + info)
            ))),
        ]


# @admin.register(export_breed)
class export_breedAdmin(admin.ModelAdmin):
    def __init__(self, model, admin_site):
        super().__init__(breedingInfo, admin_site)
        self.model = breedingInfo
        self.to_zhong = {
        }
        self.basic_fields = ['pre_num', 'ele_num', 'variety', 'sex', 'purpose', 'state', 'house_name', 'hurdle_name',
                             'birth', 'bir_weight', 'mon_age']
        self.pregnant_fields = ['check_type', 'In_pregnancy']
        self.postnatal_fields = ['delivery_date', 'ewe_health', 'live_num', 'Booroola']
        self.lamb_fields = ['ele_num', 'variety', 'sex', 'bir_weight']

    actions_selection_counter = False
    field_name = breedingInfo()._meta.fields
    list_display = ['ele_id_ewe', 'pre_num_ewe', 'ele_id_ram', 'pre_num_ram']

    list_filter = []
    exclude_display = []
    exclude_filter = []
    search_fields = []

    doit(field_name, list_display, exclude_display, list_filter, exclude_filter, search_fields)
    del list_display[0]  ##删除修改操作
    change_list_template = 'e_breed/export_change_list_zhu.html'
    add_form_template = 'e_breed/test_breedinginfo.html'
    change_form_template = 'e_breed/test_breedinginfo.html'
    change_list_template_page = 'test_change_list.html'

    ##解除只能选中执行action的限制
    def changelist_view(self, request, extra_context=None):
        """
        The 'change list' admin view for this model.
        """

        ##解除只能选中执行的限制，增加是否执行act的bool
        act_do = request.POST['action'] == 'exportAll_to_excel' if 'action' in request.POST else False

        from django.contrib.admin.views.main import ERROR_FLAG
        opts = self.model._meta
        app_label = opts.app_label
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied
        try:
            cl = self.get_changelist_instance(request)
        except IncorrectLookupParameters:
            # Wacky lookup parameters were given, so redirect to the main
            # changelist page, without parameters, and pass an 'invalid=1'
            # parameter via the query string. If wacky parameters were given
            # and the 'invalid=1' parameter was already in the query string,
            # something is screwed up with the database, so display an error
            # page.
            if ERROR_FLAG in request.GET:
                return SimpleTemplateResponse('admin/invalid_setup.html', {
                    'title': _('Database error'),
                })
            return HttpResponseRedirect(request.path + '?' + ERROR_FLAG + '=1')

        # If the request was POSTed, this might be a bulk action or a bulk
        # edit. Try to look up an action or confirmation first, but if this
        # isn't an action the POST will fall through to the bulk edit check,
        # below.
        action_failed = False
        selected = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)

        actions = self.get_actions(request)
        # Actions with no confirmation
        if (actions and request.method == 'POST' and
                'index' in request.POST and '_save' not in request.POST):
            if selected or act_do:
                response = self.response_action(request, queryset=cl.get_queryset(request))
                if response:
                    return response
                else:
                    action_failed = True
            else:
                msg = _("Items must be selected in order to perform "
                        "actions on them. No items have been changed.")
                self.message_user(request, msg, messages.WARNING)
                action_failed = True

        if (actions and request.method == 'POST' and
            helpers.ACTION_CHECKBOX_NAME in request.POST and
            'index' not in request.POST and '_save' not in request.POST) or act_do:
            if selected or act_do:
                response = self.response_action(request, queryset=cl.get_queryset(request))
                if response:
                    return response
                else:
                    action_failed = True

        if action_failed:
            # Redirect back to the changelist page to avoid resubmitting the
            # form if the user refreshes the browser or uses the "No, take
            # me back" button on the action confirmation page.
            return HttpResponseRedirect(request.get_full_path())

        # If we're allowing changelist editing, we need to construct a formset
        # for the changelist given all the fields to be edited. Then we'll
        # use the formset to validate/process POSTed data.
        formset = cl.formset = None

        # Handle POSTed bulk-edit data.
        if request.method == 'POST' and cl.list_editable and '_save' in request.POST:
            if not self.has_change_permission(request):
                raise PermissionDenied
            FormSet = self.get_changelist_formset(request)
            modified_objects = self._get_list_editable_queryset(request, FormSet.get_default_prefix())
            formset = cl.formset = FormSet(request.POST, request.FILES, queryset=modified_objects)
            if formset.is_valid():
                changecount = 0
                for form in formset.forms:
                    if form.has_changed():
                        obj = self.save_form(request, form, change=True)
                        self.save_model(request, obj, form, change=True)
                        self.save_related(request, form, formsets=[], change=True)
                        change_msg = self.construct_change_message(request, form, None)
                        self.log_change(request, obj, change_msg)
                        changecount += 1

                if changecount:
                    msg = ngettext(
                        "%(count)s %(name)s was changed successfully.",
                        "%(count)s %(name)s were changed successfully.",
                        changecount
                    ) % {
                              'count': changecount,
                              'name': model_ngettext(opts, changecount),
                          }
                    self.message_user(request, msg, messages.SUCCESS)

                return HttpResponseRedirect(request.get_full_path())

        # Handle GET -- construct a formset for display.
        elif cl.list_editable and self.has_change_permission(request):
            FormSet = self.get_changelist_formset(request)
            formset = cl.formset = FormSet(queryset=cl.result_list)

        # Build the list of media to be used by the formset.
        if formset:
            media = self.media + formset.media
        else:
            media = self.media

        # Build the action form and populate it with available actions.
        model_name = '_'.join(str(cl.opts).split('.'))
        if actions:
            action_form = self.action_form(auto_id=None)
            action_form.fields['action'].choices = self.get_action_choices_no_delete(request, model_name)
            media += action_form.media
        else:
            action_form = None

        selection_note_all = ngettext(
            '%(total_count)s selected',
            'All %(total_count)s selected',
            cl.result_count
        )
        context = {
            **self.admin_site.each_context(request),
            'module_name': str(opts.verbose_name_plural),
            'selection_note': _('0 of %(cnt)s selected') % {'cnt': len(cl.result_list)},
            'selection_note_all': selection_note_all % {'total_count': cl.result_count},
            'title': cl.title,
            'is_popup': cl.is_popup,
            'to_field': cl.to_field,
            'model_name': model_name,
            'cl': cl,
            'media': media,
            'has_add_permission': self.has_add_permission(request),
            'opts': cl.opts,
            'action_form': action_form,
            'actions_on_top': self.actions_on_top,
            'actions_on_bottom': self.actions_on_bottom,
            'actions_selection_counter': self.actions_selection_counter,
            'preserved_filters': self.get_preserved_filters(request),
            **(extra_context or {}),
        }

        request.current_app = self.admin_site.name

        return TemplateResponse(request, self.change_list_template or [
            'admin/%s/%s/change_list.html' % (app_label, opts.model_name),
            'admin/%s/change_list.html' % app_label,
            'admin/change_list.html'
        ], context)

    def response_action(self, request, queryset):
        """
        Handle an admin action. This is called if a request is POSTed to the
        changelist; it returns an HttpResponse if the action was handled, and
        None otherwise.
        """
        act_do = request.POST['action'] == 'exportAll_to_excel' if 'action' in request.POST else False
        # There can be multiple action forms on the page (at the top
        # and bottom of the change list, for example). Get the action
        # whose button was pushed.

        try:
            action_index = int(request.POST.get('index', 0))
        except ValueError:
            action_index = 0

        # Construct the action form.
        data = request.POST.copy()
        data.pop(helpers.ACTION_CHECKBOX_NAME, None)
        data.pop("index", None)

        # Use the action whose button was pushed
        try:
            data.update({'action': data.getlist('action')[action_index]})
        except IndexError:
            # If we didn't get an action from the chosen form that's invalid
            # POST data, so by deleting action it'll fail the validation check
            # below. So no need to do anything here
            pass

        action_form = self.action_form(data, auto_id=None)
        action_form.fields['action'].choices = self.get_action_choices(request)

        # If the form's valid we can handle the action.
        if action_form.is_valid():
            action = action_form.cleaned_data['action']
            select_across = action_form.cleaned_data['select_across']
            func = self.get_actions(request)[action][0]

            # Get the list of selected PKs. If nothing's selected, we can't
            # perform an action on it, so bail. Except we want to perform
            # the action explicitly on all objects.
            selected = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)
            if not selected and not select_across and not act_do:
                # Reminder that something needs to be selected or nothing will happen
                msg = _("Items must be selected in order to perform "
                        "actions on them. No items have been changed.")
                self.message_user(request, msg, messages.WARNING)
                return None

            if not select_across:
                # Perform the action only on the selected objects
                queryset = queryset.filter(pk__in=selected)

            response = func(self, request, queryset)

            # Actions may return an HttpResponse-like object, which will be
            # used as the response from the POST. If not, we'll be a good
            # little HTTP citizen and redirect back to the changelist page.
            if isinstance(response, HttpResponseBase):
                return response
            else:
                return HttpResponseRedirect(request.get_full_path())
        else:
            msg = _("No action selected.")
            self.message_user(request, msg, messages.WARNING)
            return None

    # 检验这个字段是否是choice字段，如果是就需要将其对应起来
    def check_choices(self, models):
        data = {}
        for i in models:
            fields = i._meta.fields
            for j in fields:
                if j.choices:
                    data[j.name] = dict(j.choices)
        return data

    def excel_header(self, fields_name):
        dict = self.to_zhong
        data = []
        # basic=[i.verbose_name for i in BasicInfo._meta.fields if i.name in self.basic_fields]# 导入基本信息字段名

        breed = [dict[i.name] if dict.get(i.name) else i.verbose_name for i in self.model._meta.fields if
                 i.name in fields_name]
        basic_all = {}
        basic = []
        pregnant_all = {}
        pregnant = []
        postnata_all = {}
        postnata = []
        lamb_all = {}
        lamb = []
        for i in BasicInfo._meta.fields:
            basic_all[i.name] = i.verbose_name
        for i in self.basic_fields:
            basic.append(basic_all[i])

        for i in pregnantInfo._meta.fields:
            pregnant_all[i.name] = i.verbose_name
        for i in self.pregnant_fields:
            pregnant.append(pregnant_all[i])

        for i in postnatalInfo._meta.fields:
            postnata_all[i.name] = i.verbose_name
        for i in self.postnatal_fields:
            postnata.append(postnata_all[i])

        for i in LambInfo._meta.fields:
            lamb_all[i.name] = i.verbose_name
        for j in range(1, 6):
            for i in self.lamb_fields:
                lamb.append('后裔' + str(j) + lamb_all[i])
        basic.extend(['父电子耳号', '母电子耳号', '祖父电子耳号', '祖母电子耳号', '外祖父电子耳号', '外祖母电子耳号'])
        data.extend(basic)
        data.extend(breed)
        data.extend(pregnant)
        data.extend(postnata)
        data.extend(lamb)
        return data

    def add_basic(self, data, id):
        basic = []
        for i in self.basic_fields:
            file = BasicInfo.objects.filter(id=id).values(i).first()
            if i in self.choice_field.keys() and file:
                basic.append(self.choice_field[i][file[i]] if file[i] in self.choice_field[i].keys() else file[i])
            else:
                basic.append(file[i] if file else '')
        f_id = BasicInfo.objects.filter(id=id).values('father_id').first()['father_id'] if BasicInfo.objects.filter(
            id=id).values('father_id').first() else 0
        m_id = BasicInfo.objects.filter(id=id).values('mother_id').first()['mother_id'] if BasicInfo.objects.filter(
            id=id).values('mother_id') else 0
        ff_id = BasicInfo.objects.filter(id=f_id).values('father_id').first()['father_id'] if BasicInfo.objects.filter(
            id=f_id).values('father_id').first() else 0
        fm_id = BasicInfo.objects.filter(id=f_id).values('mother_id').first()['mother_id'] if BasicInfo.objects.filter(
            id=f_id).values('mother_id').first() else 0
        mf_id = BasicInfo.objects.filter(id=m_id).values('father_id').first()['father_id'] if BasicInfo.objects.filter(
            id=f_id).values('father_id').first() else 0
        mm_id = BasicInfo.objects.filter(id=m_id).values('mother_id').first()['mother_id'] if BasicInfo.objects.filter(
            id=f_id).values('mother_id').first() else 0
        basic.extend([
            BasicInfo.objects.filter(id=f_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=f_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=m_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=m_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=ff_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=ff_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=fm_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=fm_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=mf_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=mf_id).values('ele_num').first() else '无',
            BasicInfo.objects.filter(id=mm_id).values('ele_num').first()['ele_num'] if BasicInfo.objects.filter(
                id=mm_id).values('ele_num').first() else '无',
        ])  ##添加爸爸妈妈爷爷奶奶姥姥姥爷

        return basic

    def add_breed(self, obj, fields_name):
        add = []
        for i in fields_name:
            file = eval('obj.' + i)
            if i in self.choice_field.keys():
                file = self.choice_field[i][file]
            add.append(file)
        return add

    def add_pregnant(self, data, id):
        basic = []
        for i in self.pregnant_fields:
            file = pregnantInfo.objects.filter(breeding_id=id).values(i).first()
            if i in self.choice_field.keys() and file:
                basic.append(self.choice_field[i][file[i]] if file[i] in self.choice_field[i].keys() else file[i])
            else:
                basic.append(file[i] if file else '')

        return basic

    def add_postnatal(self, data, id):
        basic = []
        for i in self.postnatal_fields:
            file = postnatalInfo.objects.filter(breeding_id=id).values(i).first()
            if i in self.choice_field.keys() and file:
                print(file[i])
                if file[i] == 'TRUE':
                    basic.append('正常')
                elif file[i] == 'FALSE':
                    basic.append('不正常')
                else:
                    basic.append(self.choice_field[i][file[i]] if file[i] in self.choice_field[i].keys() else file[i])
            else:
                basic.append(file[i] if file else '')
                # print(file)
                # if 'ewe_health' in file:
                #     if file[i]==True:
                #         basic.append('正常')
                #     elif file[i]==False:
                #         basic.append('不正常')
                # else:

        return basic

    def add_lamb(self, data, id):
        basic = []
        ids = LambInfo.objects.filter(breeding_id=id).values('id')
        for l_id in ids:
            for i in self.lamb_fields:
                file = LambInfo.objects.filter(id=l_id['id']).values(i).first()
                if i in self.choice_field.keys() and file:
                    print('yes')
                    print(file)
                    basic.append(self.choice_field[i][file[i]] if file[i] in self.choice_field[i].keys() else file[i])
                else:
                    basic.append(file[i] if file else '')
        print(basic)
        return basic

    def data(self, obj, fields_name):
        self.choice_field = self.check_choices([breedingInfo(), BasicInfo()])
        data = []
        id = getattr(obj, 'id')
        ewe_id = getattr(obj, 'ewe_id')
        ram_id = getattr(obj, 'ram_id')
        data.extend(self.add_basic(data, ewe_id))  ##添加基本信息
        data.extend(self.add_breed(obj, fields_name))
        data.extend(self.add_pregnant(data, id))
        data.extend(self.add_postnatal(data, id))
        lamb = self.add_lamb(data, id)
        data.extend(lamb if lamb else ['', ])

        # for field in fields_name:
        #     attr=getattr(obj, field)
        #     data.append(attr)
        return data

    def export_to_excel(self, request, queryset):
        fields_name = [i.name for i in self.model._meta.fields]
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response[
            'Content-Disposition'] = f'attachment; filename={self.model._meta.verbose_name.title()}.xlsx'.encode(
            'utf-8')  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        dict = self.to_zhong

        fields_name.remove('id')  ##删除id字段
        fields_name.remove('ewe_id')
        fields_name.remove('ram_id')
        fields_name.remove('belong')
        new_fields_name = self.excel_header(fields_name)

        ws.append(new_fields_name)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            data = self.data(obj, fields_name)
            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response

    def exportAll_to_excel(self, request, queryset):
        queryset = self.model.objects.all()

        fields_name = [i.name for i in self.model._meta.fields]

        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response[
            'Content-Disposition'] = f'attachment; filename={self.model._meta.verbose_name.title()}.xlsx'.encode(
            'utf-8')  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        dict = self.to_zhong

        fields_name.remove('id')  ##删除id字段
        fields_name.remove('ewe_id')
        fields_name.remove('ram_id')
        fields_name.remove('belong')
        new_fields_name = self.excel_header(fields_name)

        ws.append(new_fields_name)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            data = self.data(obj, fields_name)

            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response

    export_to_excel.allowed_permissions = ('export',)
    export_to_excel.short_description = gettext_lazy("导出所选数据")
    exportAll_to_excel.allowed_permissions = ('exportAll',)
    exportAll_to_excel.short_description = gettext_lazy("导出所有数据")

    actions = [export_to_excel, exportAll_to_excel]
